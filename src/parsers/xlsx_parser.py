import re
import logging
from pathlib import Path
from typing import List, Optional

import openpyxl

from src.models.textbook import Chapter, Textbook
from src.parsers.base import BaseParser
from src.parsers.factory import register_parser

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):

    def parse(self, file_path: Path) -> Textbook:
        logger.info(f"Parsing XLSX: {file_path}")

        try:
            with openpyxl.load_workbook(str(file_path), read_only=True) as wb:
                chapters = self._extract_chapters(wb)

            if not chapters:
                chapters.append(Chapter(
                    chapter_id=self._generate_chapter_id(0),
                    title=self._extract_title_from_filename(file_path),
                    page_start=0,
                    page_end=0,
                    content="No content found",
                    char_count=0
                ))

            total_chars = sum(ch.char_count for ch in chapters)

            return Textbook(
                textbook_id=file_path.stem,
                filename=file_path.name,
                title=self._extract_title_from_filename(file_path),
                total_pages=0,
                total_chars=total_chars,
                chapters=chapters
            )
        except Exception as e:
            logger.error(f"Failed to parse XLSX: {e}")
            raise

    def _extract_chapters(self, wb) -> List[Chapter]:
        chapters: List[Chapter] = []
        chapter_index = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content_lines = []

            for row in ws.iter_rows(values_only=True):
                line = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                if line.strip():
                    content_lines.append(line)

            if content_lines:
                content = '\n'.join(content_lines)
                chapters.append(Chapter(
                    chapter_id=self._generate_chapter_id(chapter_index),
                    title=sheet_name,
                    page_start=0,
                    page_end=0,
                    content=content,
                    char_count=len(content)
                ))
                chapter_index += 1

        return chapters


register_parser("xlsx", XlsxParser)
